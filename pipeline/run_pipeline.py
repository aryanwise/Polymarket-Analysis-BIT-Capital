"""
pipeline/run_pipeline.py

Orchestrates the full BIT Capital ETL pipeline:
  ingest → stage1 → stage2 → report_generator

Also exports a debug Excel file with one sheet per stage.

Usage:
  python pipeline/run_pipeline.py                    # full run
  python pipeline/run_pipeline.py --dry-run          # no DB writes
  python pipeline/run_pipeline.py --skip-report      # skip report generation
  python pipeline/run_pipeline.py --max-events 500   # smaller ingest for testing
"""

import sys
import os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import argparse
import logging
import pandas as pd
from datetime import datetime, timezone
from dotenv import load_dotenv

load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)

# ── Output paths ──────────────────────────────────────────────
PIPELINE_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(PIPELINE_DIR)
DEBUG_DIR    = os.path.join(PROJECT_ROOT, "debug")
os.makedirs(DEBUG_DIR, exist_ok=True)


def export_debug_excel(
    df_raw:      pd.DataFrame,
    df_stage1:   pd.DataFrame,
    df_signals:  pd.DataFrame,
    stats_s1:    dict,
    stats_s2:    dict,
    run_ts:      str,
):
    """
    Exports all three pipeline stages into one Excel file.
    Each stage = one sheet. Includes a summary sheet.
    File name: debug/pipeline_debug_YYYYMMDD_HHMMSS.xlsx
    """
    filename = os.path.join(DEBUG_DIR, f"pipeline_debug_{run_ts}.xlsx")

    with pd.ExcelWriter(filename, engine="openpyxl") as writer:

        # ── Sheet 1: Summary ──────────────────────────────────
        summary_data = {
            "Stage":   ["Ingest (raw)", "Stage 1 (filtered)", "Stage 2 (signals)"],
            "Rows":    [len(df_raw), len(df_stage1), len(df_signals)],
            "Removed": [
                0,
                len(df_raw) - len(df_stage1),
                len(df_stage1) - len(df_signals.drop_duplicates("market_id") if not df_signals.empty else df_signals),
            ],
            "Notes": [
                "All active Polymarket markets",
                f"Removed: vol=0({stats_s1.get('zero_volume',0)}) "
                f"expired({stats_s1.get('expired',0)}) "
                f"resolved({stats_s1.get('fully_resolved',0)}) "
                f"near-certain({stats_s1.get('near_certain',0)}) "
                f"bad-tags({stats_s1.get('irrelevant_tags',0)})",
                f"Signals kept: {stats_s2.get('signals',0)} | "
                f"Noise: {stats_s2.get('noise',0)} | "
                f"Errors: {stats_s2.get('errors',0)}",
            ],
        }
        pd.DataFrame(summary_data).to_excel(
            writer, sheet_name="Summary", index=False
        )

        # ── Sheet 2: Raw ingest ───────────────────────────────
        raw_export = df_raw.copy()
        # Limit to 10k rows for Excel performance
        if len(raw_export) > 10_000:
            raw_export = raw_export.head(10_000)
            logger.info("Raw sheet capped at 10,000 rows for Excel performance")
        raw_export.to_excel(writer, sheet_name="01_Ingest_Raw", index=False)

        # ── Sheet 3: Stage 1 output ───────────────────────────
        df_stage1.to_excel(writer, sheet_name="02_Stage1_Filtered", index=False)

        # ── Sheet 4: Stage 2 signals ──────────────────────────
        if not df_signals.empty:
            df_signals.to_excel(writer, sheet_name="03_Stage2_Signals", index=False)
        else:
            pd.DataFrame({"note": ["No signals found"]}).to_excel(
                writer, sheet_name="03_Stage2_Signals", index=False
            )

        # ── Style summary sheet ───────────────────────────────
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            wb  = writer.book
            ws  = wb["Summary"]
            hdr = Font(bold=True, color="FFFFFF")
            fill = PatternFill(fill_type="solid", fgColor="1F4E79")
            for cell in ws[1]:
                cell.font      = hdr
                cell.fill      = fill
                cell.alignment = Alignment(horizontal="center")
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = 40
        except Exception:
            pass   # styling is optional

    logger.info("Debug Excel saved → %s", filename)
    return filename


def run_pipeline(
    max_events:  int  = 3000,
    dry_run:     bool = False,
    skip_report: bool = False,
) -> dict:
    """
    Runs the full pipeline end-to-end.
    Returns summary dict.
    """
    run_start = datetime.now(timezone.utc)
    run_ts    = run_start.strftime("%Y%m%d_%H%M")

    logger.info("=" * 60)
    logger.info("BIT CAPITAL PIPELINE START — %s", run_ts)
    logger.info("dry_run=%s | max_events=%d | skip_report=%s", dry_run, max_events, skip_report)
    logger.info("=" * 60)

    # ── Supabase client (service role for writes) ─────────────
    supabase = None
    if not dry_run:
        try:
            from utils.supabase_client import get_service_client
            supabase = get_service_client()
            logger.info("Supabase connected")
        except Exception as e:
            logger.error("Supabase connection failed: %s", e)
            logger.warning("Continuing in dry_run mode")
            dry_run = True

    # ── Step 1: Ingest ────────────────────────────────────────
    logger.info("\n--- STEP 1: INGEST ---")
    from ingest import run_ingest
    df_raw = run_ingest(max_events=max_events)

    if df_raw.empty:
        logger.error("Ingest returned no data. Aborting.")
        return {"status": "failed", "reason": "empty_ingest"}

    # ── Step 2: Stage 1 filter ────────────────────────────────
    logger.info("\n--- STEP 2: STAGE 1 FILTER ---")
    from stage1_filter import run_stage1
    df_s1, stats_s1 = run_stage1(df_raw)

    if df_s1.empty:
        logger.error("Stage 1 returned no data. Aborting.")
        return {"status": "failed", "reason": "empty_stage1"}

    # ── Step 3: Stage 2 filter (Gemini) ──────────────────────
    logger.info("\n--- STEP 3: STAGE 2 FILTER (GEMINI) ---")
    from stage2_filter import run_stage2
    df_signals, stats_s2 = run_stage2(df_s1, supabase=supabase, dry_run=dry_run)

    # ── Step 4: Export debug Excel ────────────────────────────
    logger.info("\n--- STEP 4: DEBUG EXPORT ---")
    debug_file = export_debug_excel(
        df_raw, df_s1, df_signals,
        stats_s1, stats_s2, run_ts,
    )

    # ── Step 5: Report generator ──────────────────────────────
    report_id = None
    if not skip_report and not dry_run and not df_signals.empty:
        logger.info("\n--- STEP 5: REPORT GENERATOR ---")
        try:
            from report_generator import run_report_pipeline
            result = run_report_pipeline()
            if result:
                report_id = result.get("report_id")
                logger.info("Report generated — ID: %s", report_id)
        except Exception as e:
            logger.error("Report generation failed: %s", e)
    elif skip_report:
        logger.info("\n--- STEP 5: REPORT skipped ---")

    # ── Final summary ─────────────────────────────────────────
    elapsed = (datetime.now(timezone.utc) - run_start).total_seconds()

    summary = {
        "status":          "success",
        "run_ts":          run_ts,
        "duration_s":      round(elapsed, 1),
        "ingest_rows":     len(df_raw),
        "stage1_rows":     len(df_s1),
        "stage2_signals":  len(df_signals),
        "db_writes":       stats_s2.get("db_writes", 0),
        "report_id":       report_id,
        "debug_file":      debug_file,
        "stage1_stats":    stats_s1,
        "stage2_stats":    stats_s2,
    }

    logger.info("=" * 60)
    logger.info("PIPELINE COMPLETE — %.1fs", elapsed)
    logger.info("  Ingest rows    : %d", summary["ingest_rows"])
    logger.info("  After Stage 1  : %d  (%.1f%% kept)", summary["stage1_rows"], stats_s1.get("keep_pct", 0))
    logger.info("  Signals (S2)   : %d", summary["stage2_signals"])
    logger.info("  DB writes      : %d", summary["db_writes"])
    logger.info("  Report ID      : %s", summary["report_id"])
    logger.info("  Debug file     : %s", summary["debug_file"])
    logger.info("=" * 60)

    return summary


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="BIT Capital Pipeline")
    parser.add_argument("--max-events",  type=int,  default=3000)
    parser.add_argument("--dry-run",     action="store_true", help="No DB writes")
    parser.add_argument("--skip-report", action="store_true", help="Skip report generation")
    args = parser.parse_args()

    result = run_pipeline(
        max_events=args.max_events,
        dry_run=args.dry_run,
        skip_report=args.skip_report,
    )
    print("\nPipeline result:", result["status"])
    print("Debug file:", result.get("debug_file"))